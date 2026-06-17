"""
=====================================================================
  WHATSAPP GROUP & DM MAPPING CONFIGURATION
  
  Maps Market/District combinations to:
  - WhatsApp group names
  - Send modes (tag all DMs vs send to group)
  - DM names to tag/mention
=====================================================================
"""

import openpyxl
import json
import re

# ─────────────────────────────────────────────────────────────
#  WHATSAPP GROUP MAPPING
# ─────────────────────────────────────────────────────────────

WHATSAPP_GROUP_MAP = {
    # ─────────────────────────────────────────────────────────────
    # ARIZONA - All districts in one group, tag all DMs
    # ─────────────────────────────────────────────────────────────
    ("ARIZONA", "East"):                {"group": "ARIZONA SUPPORT", "mode": "tag_all_dms", "merge_key": "ARIZONA_ALL"},
    ("ARIZONA", "North"):               {"group": "ARIZONA SUPPORT", "mode": "tag_all_dms", "merge_key": "ARIZONA_ALL"},
    ("ARIZONA", "Central"):             {"group": "ARIZONA SUPPORT", "mode": "tag_all_dms", "merge_key": "ARIZONA_ALL"},
    ("ARIZONA", "West"):                {"group": "ARIZONA SUPPORT", "mode": "tag_all_dms", "merge_key": "ARIZONA_ALL"},
    ("ARIZONA", "Central /South Valley"):{"group": "ARIZONA SUPPORT", "mode": "tag_all_dms", "merge_key": "ARIZONA_ALL"},
    ("ARIZONA", "East Valley"):         {"group": "ARIZONA SUPPORT", "mode": "tag_all_dms", "merge_key": "ARIZONA_ALL"},
    ("ARIZONA", "West Valley"):         {"group": "ARIZONA SUPPORT", "mode": "tag_all_dms", "merge_key": "ARIZONA_ALL"},
    ("ARIZONA", ""):                    {"group": "ARIZONA SUPPORT", "mode": "tag_all_dms", "merge_key": "ARIZONA_ALL"},
    ("ARIZONA", None):                  {"group": "ARIZONA SUPPORT", "mode": "tag_all_dms", "merge_key": "ARIZONA_ALL"},
    
    # ─────────────────────────────────────────────────────────────
    # HOUSTON - Different groups per district
    # ─────────────────────────────────────────────────────────────
    ("HOUSTON", "South"):               {"group": "SOUTH DISTRICT - HOUSTON", "mode": "send_to_group"},
    ("HOUSTON", "Central"):             {"group": "CENTRAL DISTRICT - HOUSTON", "mode": "send_to_group"},
    ("HOUSTON", "North"):               {"group": "NORTH DISTRICT - HOUSTON", "mode": "send_to_group"},
    ("HOUSTON", "Airline"):             {"group": "AIRLINE DISTRICT - HOUSTON", "mode": "send_to_group"},
    ("HOUSTON", "East"):                {"group": "EAST DISTRICT - HOUSTON", "mode": "send_to_group"},
    ("HOUSTON", ""):                    {"group": "SOUTH DISTRICT - HOUSTON", "mode": "send_to_group"},
    ("HOUSTON", None):                  {"group": "SOUTH DISTRICT - HOUSTON", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # LOS ANGELES - Different groups per district
    # ─────────────────────────────────────────────────────────────
    ("LOS ANGELES", "LA - East"):       {"group": "LOS ANGELES - EAST", "mode": "send_to_group"},
    ("LOS ANGELES", "LA North"):        {"group": "LOS ANGELES - NORTH", "mode": "send_to_group"},
    ("LOS ANGELES", "San Bernardino"):  {"group": "LOS ANGELES - San Bernardino", "mode": "send_to_group"},
    ("LOS ANGELES", "LA - Central"):    {"group": "LOS ANGELES - CENTRAL", "mode": "send_to_group"},
    ("LOS ANGELES", "EAST"):            {"group": "LOS ANGELES - EAST", "mode": "send_to_group"},
    ("LOS ANGELES", "North"):           {"group": "LOS ANGELES - NORTH", "mode": "send_to_group"},
    ("LOS ANGELES", "san Bernardino"):  {"group": "LOS ANGELES - San Bernardino", "mode": "send_to_group"},
    ("LOS ANGELES", "CENTRAL"):         {"group": "LOS ANGELES - CENTRAL", "mode": "send_to_group"},
    ("LOS ANGELES", ""):                {"group": "LOS ANGELES - CENTRAL", "mode": "send_to_group"},
    ("LOS ANGELES", None):              {"group": "LOS ANGELES - CENTRAL", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # MEMPHIS - Multiple groups
    # ─────────────────────────────────────────────────────────────
    ("MEMPHIS", "Central"):             {"group": "Memphis Team Central", "mode": "send_to_group"},
    ("MEMPHIS", "South"):               {"group": "Memphis Team South", "mode": "send_to_group"},
    ("MEMPHIS", "North"):               {"group": "Memphis Team North", "mode": "send_to_group"},
    ("MEMPHIS", ""):                    {"group": "Memphis Team Central", "mode": "send_to_group"},
    ("MEMPHIS", None):                  {"group": "Memphis Team Central", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # ARKANSAS
    # ─────────────────────────────────────────────────────────────
    ("ARKANSAS", "Central"):            {"group": "Memphis Team Central", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # BAY AREA - Multiple groups
    # ─────────────────────────────────────────────────────────────
    ("BAY AREA", ""):                   {"group": "Bay Area Core", "mode": "send_to_group"},
    ("BAY AREA", None):                 {"group": "Bay Area Core", "mode": "send_to_group"},
    ("BAY AREA", "-"):                  {"group": "Bay Area Core", "mode": "send_to_group"},
    
    ("EAST BAY AREA", ""):              {"group": "East Bay Area", "mode": "send_to_group"},
    ("EAST BAY AREA", None):            {"group": "East Bay Area", "mode": "send_to_group"},

    
    ("NORTH BAY AREA", ""):             {"group": "North Bay Area- Main", "mode": "send_to_group"},
    ("NORTH BAY AREA", None):           {"group": "North Bay Area- Main", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # BOSTON
    # ─────────────────────────────────────────────────────────────
    ("BOSTON", ""):                     {"group": "Boston-Maine", "mode": "send_to_group"},
    ("BOSTON", None):                   {"group": "Boston-Maine", "mode": "send_to_group"},
    ("BOSTON", "-"):                    {"group": "Boston-Maine", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # CHARLOTTE
    # ─────────────────────────────────────────────────────────────
    ("CHARLOTTE", ""):                  {"group": "Charlotte-Support", "mode": "send_to_group"},
    ("CHARLOTTE", None):                {"group": "Charlotte-Support", "mode": "send_to_group"},
    ("CHARLOTTE", "-"):                 {"group": "Charlotte-Support", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # COLORADO - Northside Denver and Southside Denver
    # ─────────────────────────────────────────────────────────────
    ("COLORADO", "North"):              {"group": "Northside Denver", "mode": "send_to_group"},
    ("COLORADO", "South"):              {"group": "Southside Denver", "mode": "send_to_group"},
    ("COLORADO", "Southside Denver"):   {"group": "Southside Denver", "mode": "send_to_group"},
    ("COLORADO", "Northside Denver"):   {"group": "Northside Denver", "mode": "send_to_group"},
    ("COLORADO", ""):                   {"group": "Southside Denver", "mode": "send_to_group"},
    ("COLORADO", None):                 {"group": "Southside Denver", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # DALLAS - BOTH groups get complete data
    # ─────────────────────────────────────────────────────────────
    ("DALLAS", ""):                     {"group": "DALLAS_BOTH", "mode": "send_to_both_groups", "merge_key": "DALLAS_ALL"},
    ("DALLAS", None):                   {"group": "DALLAS_BOTH", "mode": "send_to_both_groups", "merge_key": "DALLAS_ALL"},
    ("DALLAS", "-"):                    {"group": "DALLAS_BOTH", "mode": "send_to_both_groups", "merge_key": "DALLAS_ALL"},
    
    # ─────────────────────────────────────────────────────────────
    # EL PASO
    # ─────────────────────────────────────────────────────────────
    ("EL PASO", ""):                    {"group": "EL PASO-LAS CRUCES MAIN", "mode": "send_to_group"},
    ("EL PASO", None):                  {"group": "EL PASO-LAS CRUCES MAIN", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # FLORIDA, KENTUCKY, NASHVILLE - Shared group
    # ─────────────────────────────────────────────────────────────
    ("FLORIDA", ""):                    {"group": "NASHVILLE - SUPPORT", "mode": "send_to_group", "merge_key": "NASHVILLE_REGION"},
    ("FLORIDA", None):                  {"group": "NASHVILLE - SUPPORT", "mode": "send_to_group", "merge_key": "NASHVILLE_REGION"},
    ("KENTUCKY", ""):                   {"group": "NASHVILLE - SUPPORT", "mode": "send_to_group", "merge_key": "NASHVILLE_REGION"},
    ("KENTUCKY", None):                 {"group": "NASHVILLE - SUPPORT", "mode": "send_to_group", "merge_key": "NASHVILLE_REGION"},
    ("NASHVILLE", ""):                  {"group": "NASHVILLE - SUPPORT", "mode": "send_to_group", "merge_key": "NASHVILLE_REGION"},
    ("NASHVILLE", None):                {"group": "NASHVILLE - SUPPORT", "mode": "send_to_group", "merge_key": "NASHVILLE_REGION"},
    ("NASHVILLE", "-"):                 {"group": "NASHVILLE - SUPPORT", "mode": "send_to_group", "merge_key": "NASHVILLE_REGION"},

    # ─────────────────────────────────────────────────────────────
    # GEORGIA - ATLANTA MAIN
    # ─────────────────────────────────────────────────────────────
    ("GEORGIA", ""):                    {"group": "ATLANTA TEAM", "mode": "send_to_group"},
    ("GEORGIA", None):                  {"group": "ATLANTA TEAM", "mode": "send_to_group"},
    ("GEORGIA", "-"):                   {"group": "ATLANTA TEAM", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # NORTH CAROLINA - Multiple groups
    # ─────────────────────────────────────────────────────────────
    ("NORTH CAROLINA", "Durham"):       {"group": "NC - Durham", "mode": "send_to_group"},
    ("NORTH CAROLINA", "Raleigh West"): {"group": "NC - Raleigh West", "mode": "send_to_group"},
    ("NORTH CAROLINA", "Raleigh East"): {"group": "NC - Raleigh East", "mode": "send_to_group"},
    ("NORTH CAROL", "Durham"):          {"group": "NC - Durham", "mode": "send_to_group"},
    ("NORTH CAROL", "Raleigh West"):    {"group": "NC - Raleigh West", "mode": "send_to_group"},
    ("NORTH CAROL", "Raleigh East"):    {"group": "NC - Raleigh East", "mode": "send_to_group"},
    ("NORTH CAROL", ""):                {"group": "NC - Durham", "mode": "send_to_group"},
    ("NORTH CAROL", None):              {"group": "NC - Durham", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # OKLAHOMA / OKHLAHOMA (handles typo)
    # ─────────────────────────────────────────────────────────────
    ("OKLAHOMA", ""):                   {"group": "Oklahoma Core", "mode": "send_to_group"},
    ("OKLAHOMA", None):                 {"group": "Oklahoma Core", "mode": "send_to_group"},
    ("OKHLAHOMA", ""):                  {"group": "Oklahoma Core", "mode": "send_to_group"},
    ("OKHLAHOMA", None):                {"group": "Oklahoma Core", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # OREGON
    # ─────────────────────────────────────────────────────────────
    ("OREGON", ""):                     {"group": "Portland Oregon Main", "mode": "send_to_group"},
    ("OREGON", ""):                     {"group": "Portland Oregon Main", "mode": "send_to_group"},
    ("OREGON", None):                   {"group": "Portland Oregon Main", "mode": "send_to_group"},
    ("OREGON", "-"):                    {"group": "Portland Oregon Main", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # PALMDALE/OXNARD - Both go to Portland Oregon Main
    # ─────────────────────────────────────────────────────────────
    ("PALMDALE", ""):                   {"group": "Team Oxnard - Palmdale Core", "mode": "send_to_group"},
    ("PALMDALE", None):                 {"group": "Team Oxnard - Palmdale Core", "mode": "send_to_group"},
    ("OXNARD", ""):                     {"group": "Team Oxnard - Palmdale Core", "mode": "send_to_group"},
    ("OXNARD", None):                   {"group": "Team Oxnard - Palmdale Core", "mode": "send_to_group"},
    ("OXNARD", "Oxnard/Palmdale"):      {"group": "Team Oxnard - Palmdale Core", "mode": "send_to_group"},
    ("OXNARD/PALMDALE", ""):            {"group": "Team Oxnard - Palmdale Core", "mode": "send_to_group"},
    ("OXNARD/PALMDALE", None):          {"group": "Team Oxnard - Palmdale Core", "mode": "send_to_group"},
    ("OXNARD/PALMDALE", "Oxnard/Palmdale"): {"group": "Team Oxnard - Palmdale Core", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # PALM BEACH
    # ─────────────────────────────────────────────────────────────
    ("PALM BEACH", ""):                 {"group": "MIAMI - Support", "mode": "send_to_group"},
    ("PALM BEACH", None):               {"group": "MIAMI - Support", "mode": "send_to_group"},
    ("PALM BEACH", "-"):                {"group": "MIAMI - Support", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # PHILADELPHIA / PHILLY / PHILY (handles typos)
    # ─────────────────────────────────────────────────────────────
    ("PHILADELPHIA", ""):               {"group": "Philadelphia Support", "mode": "send_to_group"},
    ("PHILADELPHIA", None):             {"group": "Philadelphia Support", "mode": "send_to_group"},
    ("PHILLY", ""):                     {"group": "Philadelphia Support", "mode": "send_to_group"},
    ("PHILLY", None):                   {"group": "Philadelphia Support", "mode": "send_to_group"},
    ("PHILLY", "-"):                    {"group": "Philadelphia Support", "mode": "send_to_group"},
    ("PHILY", ""):                      {"group": "Philadelphia Support", "mode": "send_to_group"},
    ("PHILY", None):                    {"group": "Philadelphia Support", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # SACRAMENTO - All DMs in single group
    # ─────────────────────────────────────────────────────────────
    ("SACRAMENTO", ""):                 {"group": "Sacramento Main", "mode": "send_to_group", "merge_key": "SACRAMENTO_ALL"},
    ("SACRAMENTO", None):               {"group": "Sacramento Main", "mode": "send_to_group", "merge_key": "SACRAMENTO_ALL"},
    ("SACRAMENTO", "Stockton & Manteca"): {"group": "Sacramento Main", "mode": "send_to_group", "merge_key": "SACRAMENTO_ALL"},
    ("SACRAMENTO", "Merced & Modesto"): {"group": "Sacramento Main", "mode": "send_to_group", "merge_key": "SACRAMENTO_ALL"},
    
    # ─────────────────────────────────────────────────────────────
    # SAN DIEGO
    # ─────────────────────────────────────────────────────────────
    ("SAN DIEGO", ""):                  {"group": "SAN DIEGO CORE", "mode": "send_to_group"},
    ("SAN DIEGO", None):                {"group": "SAN DIEGO CORE", "mode": "send_to_group"},
    ("SAN DIEGO", "-"):                 {"group": "SAN DIEGO CORE", "mode": "send_to_group"},
    
    # ─────────────────────────────────────────────────────────────
    # SAN FRANCISCO - Shares SAN DIEGO CORE
    # ─────────────────────────────────────────────────────────────
    ("SAN FRANCISCO", ""):              {"group": "SAN DIEGO CORE", "mode": "send_to_group", "merge_key": "SAN_DIEGO_REGION"},
    ("SAN FRANCISCO", None):            {"group": "SAN DIEGO CORE", "mode": "send_to_group", "merge_key": "SAN_DIEGO_REGION"},
    ("SAN FRANCISCO", "-"):             {"group": "SAN DIEGO CORE", "mode": "send_to_group", "merge_key": "SAN_DIEGO_REGION"},
    
    # ─────────────────────────────────────────────────────────────
    # UTAH
    # ─────────────────────────────────────────────────────────────
    ("UTAH", ""):                       {"group": "Utah Support", "mode": "send_to_group"},
    ("UTAH", None):                     {"group": "Utah Support", "mode": "send_to_group"},
    
   
    ("NASHVILLE", ""): {"group": "NASHVILLE - SUPPORT", "mode": "send_to_group", "merge_key": "NASHVILLE_REGION"},
    ("NASHVILLE", None): {"group": "NASHVILLE - SUPPORT", "mode": "send_to_group", "merge_key": "NASHVILLE_REGION"},
}

# ─────────────────────────────────────────────────────────────
#  SPECIAL GROUP CONFIGS
# ─────────────────────────────────────────────────────────────

SPECIAL_GROUP_CONFIGS = {
    "ARIZONA_ALL": {
        "groups": ["ARIZONA SUPPORT"],
        "mode": "tag_all_dms",
        "description": "All Arizona DMs tagged in single reminder"
    },
    "DALLAS_ALL": {
        "groups": ["Dallas Team South", "Dallas Team North"],
        "mode": "send_to_both_groups",
        "description": "Complete data sent to both Dallas groups"
    },
    "NASHVILLE_REGION": {
        "groups": ["NASHVILLE - SUPPORT"],
        "mode": "send_to_group",
        "description": "Shared group for Nashville region"
    },
    "SAN_DIEGO_REGION": {
        "groups": ["SAN DIEGO CORE"],
        "mode": "send_to_group",
        "description": "Shared group for San Diego region"
    },
    "SACRAMENTO_ALL": {
        "groups": ["Sacramento Main"],
        "mode": "send_to_group",
        "description": "All Sacramento DMs in single reminder"
    }
}

# ─────────────────────────────────────────────────────────────
#  HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────

def normalize_name(name):
    """Normalize name for matching - remove special chars, extra spaces"""
    if not name:
        return ""
    name = re.sub(r'[^a-zA-Z0-9 ]', '', name)
    name = ' '.join(name.split())
    return name.lower().strip()

def get_group_for_market(market, district):
    """
    Get WhatsApp group info for a given market/district.
    Handles various district formats (None, "", normalized names).
    """
    # Normalize district
    if not district:
        district = ""
    district = str(district).strip()
    
    # Try exact match
    key = (market.upper(), district)
    if key in WHATSAPP_GROUP_MAP:
        return WHATSAPP_GROUP_MAP[key]
    
    # Try with empty district
    key = (market.upper(), "")
    if key in WHATSAPP_GROUP_MAP:
        return WHATSAPP_GROUP_MAP[key]
    
    # Not found
    return None

def load_dm_contacts(dm_contacts_file):
    """
    Load DM contact information from Excel file with District column.
    Returns dict: {market: {district: {dm_name: {'phone': phone, 'dmNameRep': rep_name}}}}
    """
    wb = openpyxl.load_workbook(dm_contacts_file)
    ws = wb.active
    
    dm_contacts = {}
    
    for row in range(2, ws.max_row + 1):
        market = str(ws.cell(row, 1).value or "").strip()
        district = str(ws.cell(row, 2).value or "").strip()
        name = str(ws.cell(row, 3).value or "").strip()
        phone = str(ws.cell(row, 7).value or "").strip()
        dm_name_rep = str(ws.cell(row, 8).value or "").strip()
        
        if not market or not name:
            continue
        
        if market not in dm_contacts:
            dm_contacts[market] = {}
        
        if district not in dm_contacts[market]:
            dm_contacts[market][district] = {}
        
        # Clean the values
        if phone == '-' or phone == '' or phone == 'None':
            phone = None
        
        if dm_name_rep == '-' or dm_name_rep == '' or dm_name_rep == 'None':
            dm_name_rep = None
        
        dm_contacts[market][district][name] = {
            'phone': phone,
            'dmNameRep': dm_name_rep
        }
    
    return dm_contacts

def get_dms_for_group(market, district, dm_contacts):
    """
    Get list of DMs that belong to a specific market/district.
    Returns list of dict with name, phone, and dmNameRep.
    """
    result = []
    
    if market not in dm_contacts:
        return result
    
    # If district is specified and exists, get DMs for that district only
    if district and district in dm_contacts[market]:
        dms_in_district = dm_contacts[market][district]
        for dm_name, dm_info in dms_in_district.items():
            result.append({
                'name': dm_name,
                'phone': dm_info.get('phone', ''),
                'dmNameRep': dm_info.get('dmNameRep', '')
            })
    else:
        # If no district specified, get all DMs in this market (all districts)
        for district_name, dms_in_district in dm_contacts[market].items():
            for dm_name, dm_info in dms_in_district.items():
                result.append({
                    'name': dm_name,
                    'phone': dm_info.get('phone', ''),
                    'dmNameRep': dm_info.get('dmNameRep', '')
                })
    
    return result


if __name__ == "__main__":
    # Test the config
    print("=" * 60)
    print("TESTING WHATSAPP GROUP MAPPING")
    print("=" * 60)
    
    test_cases = [
        ("ARIZONA", "East"),
        ("ARIZONA", "Central"),
        ("HOUSTON", "South"),
        ("HOUSTON", "Airline"),
        ("DALLAS", ""),
        ("MEMPHIS", "North"),
        ("COLORADO", "North"),
        ("COLORADO", "South"),
        ("SACRAMENTO", "Stockton & Manteca"),
        ("SACRAMENTO", "Merced & Modesto"),
    ]
    
    print("\nChecking all markets...")
    all_found = True
    
    for market, district in test_cases:
        result = get_group_for_market(market, district)
        status = "✅" if result else "❌ MISSING"
        print(f"{status} {market:20} | {district if district else '(empty)':20} → {result.get('group', 'NOT FOUND') if result else 'NOT FOUND'}")
        if not result:
            all_found = False
    
    print("\n" + "=" * 60)
    if all_found:
        print("✅ ALL MARKETS MAPPED SUCCESSFULLY!")
    else:
        print("❌ SOME MARKETS ARE MISSING - FIX THEM!")
    print("=" * 60)