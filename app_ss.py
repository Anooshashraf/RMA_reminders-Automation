"""
=====================================================================
  UNIVERSAL EXCEL → WHATSAPP SCREENSHOT GENERATOR
  
  TWO MODES:
  - RMA: One screenshot per WhatsApp group (all DMs combined)
  - Accessories: One screenshot per DM (separate SS for each DM)
  
  All DMs in same market → same WhatsApp group with @mentions
=====================================================================
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import win32com.client
import os
import time
from PIL import ImageGrab
import json
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from CONFIG_WhatsApp_Groups import get_group_for_market, load_dm_contacts
    HAS_CONFIG = True
except ImportError:
    HAS_CONFIG = False

# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION PRESETS
# ═══════════════════════════════════════════════════════════════

PRESETS = {
    "RMA / XBM / Trade-in": {
        "excel_file": r"C:\Users\admin\Desktop\RMA XBM Trade-in Consolidated2026.xlsx",
        "sheet_name": "Sheet1",
        "header_row": 1,
        "is_pivot": False,
        "mode": "group",  # One SS per WhatsApp group
        "columns": {
            "market": "A",
            "district": "B",
            "dm_name": "C",
            "store_id": "D",
            "store_name": "E",
            "description": "H",
            "imei": "I",
            "employee": "J",
            "assurant": "K",
            "processed_date": "L",
            "label_type": "M",
            "rma_number": "N",
            "rma_date": "O",
            "count": "P",
            "tracking": "Q",
            "status": "R",
        },
        "output_folder": "RMA_Screenshots",
    },
    "Accessories / Orders": {
        "excel_file": "",
        "sheet_name": "Sheet1",
        "header_row": 1,
        "is_pivot": True,
        "mode": "dm",
        "columns": {
            "market": "A",
            "dm_name": "B",       # ← Column B (no district!)
            "store_name": "C",    # ← Column C
            "order_no": "D",      # ← Column D
            "sku": "E",           # ← Column E
            "sku_desc": "F",      # ← Column F
            "status": "G",        # ← Column G
            "age": "H",           # ← Column H
            "count": "I",         # ← Column I
        },
        "output_folder": "Accessories_Screenshots",
    },
}

SETTINGS_FILE = "screenshot_generator_settings.json"


def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r') as f:
                return json.load(f)
        except:
            pass
    return {"presets": PRESETS, "last_used": "RMA / XBM / Trade-in"}


def save_settings(settings):
    with open(SETTINGS_FILE, 'w') as f:
        json.dump(settings, f, indent=2)


def col_to_num(col_letter):
    result = 0
    for char in col_letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def safe_filename(name):
    return "".join(c for c in name if c.isalnum() or c in " _-").strip()


def load_dm_contacts_mapping(dm_contacts_file):
    """Load DM contacts for dmNameRep mapping from Excel"""
    import openpyxl
    
    if not os.path.exists(dm_contacts_file):
        print(f"DEBUG: Contacts file not found at {dm_contacts_file}")
        return {}
    
    wb = openpyxl.load_workbook(dm_contacts_file)
    ws = wb.active
    
    dm_map = {}
    
    for row in range(2, ws.max_row + 1):
        market = str(ws.cell(row, 1).value or "").strip()
        district = str(ws.cell(row, 2).value or "").strip()
        name = str(ws.cell(row, 3).value or "").strip()
        phone = str(ws.cell(row, 7).value or "").strip()
        dm_name_rep = str(ws.cell(row, 8).value or "").strip()
        
        if not name:
            continue
        
        # Clean values
        if phone in ['-', '', 'None', 'null']:
            phone = None
        if dm_name_rep in ['-', '', 'None', 'null']:
            dm_name_rep = None
        
        norm = name.lower().strip()
        dm_map[norm] = {
            'phone': phone,
            'dmNameRep': dm_name_rep if dm_name_rep else name,
        }
        
        # DEBUG: Print what we loaded
        print(f"DEBUG DM CONTACT: '{name}' → rep='{dm_name_rep}' phone='{phone}'")
    
    wb.close()
    return dm_map

def get_dm_name_rep_fixed(dm_name):
    """
    HARDCODED DM NAME MAPPING - Guaranteed to work
    This maps full DM names to their short @mention names
    """
    MAPPING = {
        "Hamed Ali": "HAMED",
        "Hamza Ali": "hamza",
        "Akbar Uddin": "Adil",
        "Shoeb Naqvi": "Naqvi",
        "Abdullah Butt": "Abdullah",
        "Muhammad Farhan Asghar": "Farhan",
        "Talha Qureshi": "Tee",
        "Hassan Saleem": "Hassan",
        "Kamaran Mohammed": "Kamran",
        "SyedAli AhmedRizvi": "Ali",
        "Muhammad Sumairuddin": "Sumair",
        "Imran Ahmed Mohammed": "Ahmed",
        "Imran Shaikh": "Mazhar",
        "Salim Thanawala": "Salim",
        "Ayyan Budwani": "Ayan",
        "Muhammad Shoaib Sheeraz": "Shoaib",
        "MD Sunvee Bin Islam": "Sunvee",
        "Mohammed Anas": "Anas",
        "MUHAMMAD AFZAL": "MA",
        "Namir Elmouchantaf": "Namir",
        "Mukram Shareef Mohammed": "Mukarram",
        "SALMAN RIAZ": "SR",
        "Wajahat Ali Sattar Rajper": "Wajahat",
        "ZUBAIR HUSSAIN": "Zubair",
        "ABDUL RAFAY ASHRAF": "Abdul",
        "MAAZ KHAN": "Maaz",
        "SHARIK THOBANI": "Sharik",
        "SUBHAN ANSARI": "Subhan",
        "SHOAIB SHEERAZ": "Shoaib",
        "ZAID WASEEM": "Zaid",
        "SYED AMIR": "Amir",
        "Khaja Ameenuddin Ghori": "Mr.Anas",
        "MOHAMMED SUMAIR": "Sumair",
        "HASSAN TANVEER": "Hassan",
        "UZAIR UDDIN": "Uzair",
        "Prabhakar Sivan": "Prabha",
        "ASLAM KHAN": "Aslam",
        "Shahrukh Khalid": "SRK",
        "HAFIZ ASAD BURGEES": "Asad",
        "Nur Rahman": "Nur",
        "Shoeb Naqvi": "SHOEB NAQVI",
        "Haroon Iqbal": "Haroon",
        "Saad Ali": "Saad",
    }
    
    # Try exact match first
    if dm_name in MAPPING:
        return MAPPING[dm_name]
    
    # Try case-insensitive match
    for key, value in MAPPING.items():
        if key.lower() == dm_name.lower():
            return value
    
    # Try partial match
    dm_lower = dm_name.lower()
    for key, value in MAPPING.items():
        if dm_lower in key.lower() or key.lower() in dm_lower:
            return value
    
    # Fallback
    return dm_name


def find_dm_info(dm_name, dm_map):
    """Find DM info (phone, dmNameRep) from contacts map"""
    if not dm_name or not dm_map:
        print(f"DEBUG: find_dm_info fallback for '{dm_name}' (no map or empty name)")
        return {'phone': None, 'dmNameRep': dm_name}
    
    norm = dm_name.lower().strip()
    
    # 1. EXACT MATCH
    if norm in dm_map:
        info = dm_map[norm]
        print(f"DEBUG: EXACT match '{dm_name}' → rep='{info['dmNameRep']}'")
        return info
    
    # 2. CASE-INSENSITIVE MATCH
    for contact_name, info in dm_map.items():
        if contact_name.lower() == norm:
            print(f"DEBUG: CASE match '{dm_name}' → '{contact_name}' → rep='{info['dmNameRep']}'")
            return info
    
    # 3. PARTIAL MATCH
    for contact_name, info in dm_map.items():
        if norm in contact_name or contact_name in norm:
            print(f"DEBUG: PARTIAL match '{dm_name}' → '{contact_name}' → rep='{info['dmNameRep']}'")
            return info
    
    print(f"DEBUG: NO match for '{dm_name}', using full name")
    return {'phone': None, 'dmNameRep': dm_name}

class ScreenshotGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Universal Excel → WhatsApp Screenshot Generator")
        self.root.geometry("1000x750")
        
        self.settings = load_settings()
        self.current_preset = self.settings.get("last_used", "RMA / XBM / Trade-in")
        
        self.build_ui()
        self.load_preset(self.current_preset)
    
    def build_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        ttk.Label(main_frame, text="📊 Universal Screenshot Generator",
                  font=('Arial', 14, 'bold')).grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        # Preset selector
        preset_frame = ttk.LabelFrame(main_frame, text="Dataset Preset", padding="10")
        preset_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.preset_var = tk.StringVar(value=self.current_preset)
        preset_combo = ttk.Combobox(preset_frame, textvariable=self.preset_var,
                                     values=list(PRESETS.keys()), state="readonly", width=30)
        preset_combo.grid(row=0, column=0, padx=5)
        ttk.Button(preset_frame, text="Load Preset",
                   command=lambda: self.load_preset(self.preset_var.get())).grid(row=0, column=1, padx=5)
        ttk.Button(preset_frame, text="Save as New Preset",
                   command=self.save_as_preset).grid(row=0, column=2, padx=5)
        
        # Excel file
        file_frame = ttk.LabelFrame(main_frame, text="Excel File Settings", padding="10")
        file_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(file_frame, text="Excel File:").grid(row=0, column=0, sticky=tk.W)
        self.excel_file_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.excel_file_var, width=60).grid(row=0, column=1, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_excel).grid(row=0, column=2)
        
        ttk.Label(file_frame, text="Sheet Name:").grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        self.sheet_var = tk.StringVar(value="Sheet1")
        ttk.Entry(file_frame, textvariable=self.sheet_var, width=20).grid(row=1, column=1, sticky=tk.W, padx=5, pady=(5, 0))
        
        ttk.Label(file_frame, text="Header Row:").grid(row=1, column=1, sticky=tk.E, pady=(5, 0))
        self.header_var = tk.IntVar(value=1)
        ttk.Spinbox(file_frame, from_=1, to=10, textvariable=self.header_var, width=5).grid(row=1, column=2, sticky=tk.W, pady=(5, 0))
        
        # Options
        opt_frame = ttk.Frame(file_frame)
        opt_frame.grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=(5, 0))
        
        self.is_pivot_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_frame, text="Pivot Table",
                        variable=self.is_pivot_var).pack(side=tk.LEFT, padx=5)
        
        self.mode_var = tk.StringVar(value="group")
        ttk.Radiobutton(opt_frame, text="Per Group", variable=self.mode_var, 
                        value="group").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(opt_frame, text="Per DM", variable=self.mode_var, 
                        value="dm").pack(side=tk.LEFT, padx=5)
        
        # Column mapping
        col_frame = ttk.LabelFrame(main_frame, text="Column Mapping (Enter column letters: A, B, C...)", padding="10")
        col_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        main_frame.rowconfigure(3, weight=1)
        
        self.col_entries = {}
        columns_list = [
            ("Market:", "market"), ("District:", "district"), ("DM Name:", "dm_name"),
            ("Store ID:", "store_id"), ("Store Name:", "store_name"),
            ("Order No:", "order_no"), ("SKU:", "sku"), ("SKU Desc:", "sku_desc"),
            ("Description:", "description"), ("IMEI/Serial:", "imei_serial"),
            ("IMEI:", "imei"), ("Employee:", "employee"), ("Assurant:", "assurant"),
            ("Processed Date:", "processed_date"), ("Label Type:", "label_type"),
            ("RMA Number:", "rma_number"), ("RMA Date:", "rma_date"),
            ("Count:", "count"), ("Tracking:", "tracking"), ("Status:", "status"),
            ("IDOO:", "idoo"), ("AGE:", "age"),
        ]
        
        for i, (label, key) in enumerate(columns_list):
            row = i // 4
            col = (i % 4) * 2
            ttk.Label(col_frame, text=label).grid(row=row, column=col, sticky=tk.W, padx=(5, 2), pady=2)
            var = tk.StringVar()
            ttk.Entry(col_frame, textvariable=var, width=5).grid(row=row, column=col+1, sticky=tk.W, pady=2)
            self.col_entries[key] = var
        
        # Output settings
        out_frame = ttk.LabelFrame(main_frame, text="Output Settings", padding="10")
        out_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(out_frame, text="Output Folder:").grid(row=0, column=0, sticky=tk.W)
        self.output_var = tk.StringVar(value="Screenshots")
        ttk.Entry(out_frame, textvariable=self.output_var, width=30).grid(row=0, column=1, padx=5)
        
        # Control buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(10, 0))
        
        self.generate_btn = ttk.Button(btn_frame, text="🚀 GENERATE SCREENSHOTS",
                                        command=self.start_generation)
        self.generate_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="💾 Save Settings",
                   command=self.save_current).pack(side=tk.LEFT, padx=5)
        
        self.progress_var = tk.StringVar(value="Ready")
        ttk.Label(btn_frame, textvariable=self.progress_var).pack(side=tk.RIGHT, padx=10)
        
        # Log
        log_frame = ttk.LabelFrame(main_frame, text="Log", padding="5")
        log_frame.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        main_frame.rowconfigure(6, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=8, state=tk.DISABLED)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
    
    def log(self, message):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)
        self.root.update_idletasks()
    
    def browse_excel(self):
        filepath = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filepath:
            self.excel_file_var.set(filepath)
    
    def load_preset(self, preset_name):
        if preset_name not in PRESETS:
            return
        
        preset = PRESETS[preset_name]
        self.current_preset = preset_name
        
        self.excel_file_var.set(preset.get("excel_file", ""))
        self.sheet_var.set(preset.get("sheet_name", "Sheet1"))
        self.header_var.set(preset.get("header_row", 1))
        self.output_var.set(preset.get("output_folder", "Screenshots"))
        self.is_pivot_var.set(preset.get("is_pivot", False))
        self.mode_var.set(preset.get("mode", "group"))
        
        columns = preset.get("columns", {})
        for key, var in self.col_entries.items():
            var.set(columns.get(key, ""))
        
        self.log(f"✅ Loaded preset: {preset_name} | Mode: {preset.get('mode', 'group')} | Pivot: {preset.get('is_pivot', False)}")
    
    def save_as_preset(self):
        name = tk.simpledialog.askstring("Save Preset", "Enter preset name:")
        if not name:
            return
        
        columns = {}
        for key, var in self.col_entries.items():
            if var.get().strip():
                columns[key] = var.get().strip().upper()
        
        PRESETS[name] = {
            "excel_file": self.excel_file_var.get(),
            "sheet_name": self.sheet_var.get(),
            "header_row": self.header_var.get(),
            "columns": columns,
            "output_folder": self.output_var.get(),
            "is_pivot": self.is_pivot_var.get(),
            "mode": self.mode_var.get(),
        }
        
        self.settings["presets"] = PRESETS
        self.preset_var.set(name)
        self.current_preset = name
        save_settings(self.settings)
        self.log(f"💾 Saved new preset: {name}")
    
    def save_current(self):
        columns = {}
        for key, var in self.col_entries.items():
            if var.get().strip():
                columns[key] = var.get().strip().upper()
        
        if self.current_preset in PRESETS:
            PRESETS[self.current_preset].update({
                "excel_file": self.excel_file_var.get(),
                "sheet_name": self.sheet_var.get(),
                "header_row": self.header_var.get(),
                "columns": columns,
                "output_folder": self.output_var.get(),
                "is_pivot": self.is_pivot_var.get(),
                "mode": self.mode_var.get(),
            })
        
        self.settings["last_used"] = self.current_preset
        save_settings(self.settings)
        self.log("💾 Settings saved!")
    
    def start_generation(self):
        excel_file = self.excel_file_var.get()
        if not excel_file or not os.path.exists(excel_file):
            messagebox.showerror("Error", "Please select a valid Excel file!")
            return
        
        self.generate_btn.configure(state=tk.DISABLED)
        self.progress_var.set("Generating...")
        
        import threading
        thread = threading.Thread(target=self.generate_screenshots)
        thread.daemon = True
        thread.start()
    
    def generate_screenshots(self):
        try:
            excel_file = self.excel_file_var.get()
            sheet_name = self.sheet_var.get()
            header_row = self.header_var.get()
            is_pivot = self.is_pivot_var.get()
            mode = self.mode_var.get()
            output_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), self.output_var.get())
            
            col_map = {}
            for key, var in self.col_entries.items():
                if var.get().strip():
                    col_map[key] = var.get().strip().upper()
            
            if "market" not in col_map or "dm_name" not in col_map:
                self.log("❌ Missing required columns: market and dm_name")
                self.root.after(0, self.generation_done)
                return
            
            os.makedirs(output_folder, exist_ok=True)
            
            # Clear old files
            removed = 0
            for fname in os.listdir(output_folder):
                if fname.lower().endswith('.png') or fname == '_group_list.json':
                    try:
                        os.remove(os.path.join(output_folder, fname))
                        removed += 1
                    except:
                        pass
            if removed:
                self.log(f"Cleared {removed} existing files")
            
            # ═══════════════════════════════════════════════════
            #  LOAD DM CONTACTS FOR dmNameRep MAPPING
            # ═══════════════════════════════════════════════════
            
            dm_contacts_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "DM contacts.xlsx")
            dm_map = {}
            if os.path.exists(dm_contacts_file):
                try:
                    dm_map = load_dm_contacts_mapping(dm_contacts_file)
                    self.log(f"✅ Loaded {len(dm_map)} DM contacts for dmNameRep mapping")
                except Exception as e:
                    self.log(f"⚠️ Could not load DM contacts: {e}")
            else:
                self.log(f"⚠️ DM contacts.xlsx not found, using full names for tags")
            
            self.log(f"🚀 Mode: {'Per Group' if mode == 'group' else 'Per DM'} | Pivot: {is_pivot}")
            self.log("Launching Excel...")
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = True
            
            wb = excel.Workbooks.Open(excel_file)
            ws = wb.Sheets(sheet_name)
            
            col_market_num = col_to_num(col_map["market"])
            col_dm_num = col_to_num(col_map["dm_name"])
            last_row = ws.Cells(ws.Rows.Count, col_market_num).End(-4162).Row
            
            # Find last used column
            last_col = 1
            for col in range(1, 30):
                try:
                    header_val = str(ws.Cells(header_row, col).Value or "").strip()
                    data_val = str(ws.Cells(header_row + 1, col).Value or "").strip()
                    if header_val or data_val:
                        last_col = max(last_col, col)
                except:
                    pass
            
            capture_left = 1
            capture_right = last_col
            self.log(f"📐 Columns: A to {chr(64 + capture_right)}")
            
            # ═══════════════════════════════════════════════════
            #  COLLECT DATA
            # ═══════════════════════════════════════════════════
            
            market_dm_rows = {}
            
            if is_pivot:
                self.log(f"📊 Scanning pivot table ({last_row} rows)...")
                current_market = None
                current_dm = None  # ← ADD THIS: track current DM
                
                for row in range(header_row + 1, last_row + 1):
                    market_val = str(ws.Cells(row, col_market_num).Value or "").strip()
                    dm_val = str(ws.Cells(row, col_dm_num).Value or "").strip()
                    
                    if "Grand Total" in market_val or "grand total" in market_val.lower():
                        break
                    
                    if market_val and market_val != "" and market_val != "None":
                        current_market = market_val
                        if current_market not in market_dm_rows:
                            market_dm_rows[current_market] = {}
                    
                    # Track DM name (carry forward if empty)
                    if dm_val and dm_val != "" and dm_val != "None":
                        current_dm = dm_val
                    
                    # Use the tracked DM name
                    if current_market and current_dm:
                        if current_dm not in market_dm_rows[current_market]:
                            market_dm_rows[current_market][current_dm] = []
                        market_dm_rows[current_market][current_dm].append(row)
            
            # ═══════════════════════════════════════════════════
            #  GROUP BY WHATSAPP GROUP
            # ═══════════════════════════════════════════════════
            
            group_map = {}
            
            for market, dms in market_dm_rows.items():
                group_info = None
                if HAS_CONFIG:
                    group_info = get_group_for_market(market.upper(), "")
                
                if not group_info:
                    group_info = {"group": f"{market} Support", "mode": "send_to_group"}
                
                group_key = group_info.get('merge_key', group_info.get('group'))
                
                if group_key not in group_map:
                    group_map[group_key] = {
                        'whatsapp_group': group_info.get('group'),
                        'send_mode': group_info.get('mode', 'send_to_group'),
                        'markets': set(),
                        'dm_data': {},
                    }
                
                group_map[group_key]['markets'].add(market)
                
                for dm_name, rows in dms.items():
                    if dm_name not in group_map[group_key]['dm_data']:
                        group_map[group_key]['dm_data'][dm_name] = []
                    group_map[group_key]['dm_data'][dm_name].extend(rows)
            
            # ═══════════════════════════════════════════════════
            #  GENERATE SCREENSHOTS
            # ═══════════════════════════════════════════════════
            
            group_info_list = []
            ss_count = 0
            
            for group_key, group_data in sorted(group_map.items()):
                group_name = group_data['whatsapp_group']
                send_mode = group_data['send_mode']
                all_dm_names = sorted(group_data['dm_data'].keys())
                
                if mode == "group":
                    # ═══════════════════════════════════════════
                    #  RMA MODE: One SS per group
                    # ═══════════════════════════════════════════
                    
                    ss_count += 1
                    all_rows = []
                    for dm_name in all_dm_names:
                        all_rows.extend(group_data['dm_data'][dm_name])
                    all_rows.sort()
                    
                    self.log(f"[{ss_count}] {group_name} - {len(all_dm_names)} DMs, {len(all_rows)} rows")
                    self.progress_var.set(f"[{ss_count}] {group_name}")
                    
                    # Create temp sheet
                    excel.DisplayAlerts = False
                    try:
                        wb.Sheets("_TEMP_SS").Delete()
                    except:
                        pass
                    excel.DisplayAlerts = True
                    
                    excel.DisplayAlerts = False
                    temp_ws = wb.Sheets.Add()
                    temp_ws.Name = "_TEMP_SS"
                    excel.DisplayAlerts = True
                    
                    # Copy header
                    ws.Range(ws.Cells(header_row, capture_left), ws.Cells(header_row, capture_right)).Copy(
                        temp_ws.Range(temp_ws.Cells(1, 1), temp_ws.Cells(1, capture_right)))
                    temp_ws.Rows(1).RowHeight = ws.Rows(header_row).RowHeight
                    
                    # Copy all rows
                    for t_row, s_row in enumerate(all_rows, 2):
                        ws.Range(ws.Cells(s_row, capture_left), ws.Cells(s_row, capture_right)).Copy(
                            temp_ws.Range(temp_ws.Cells(t_row, 1), temp_ws.Cells(t_row, capture_right)))
                        try:
                            temp_ws.Rows(t_row).RowHeight = ws.Rows(s_row).RowHeight
                        except:
                            pass
                    
                    total_rows = len(all_rows) + 1
                    
                    for c in range(capture_left, capture_right + 1):
                        try:
                            temp_ws.Columns(c).ColumnWidth = ws.Columns(c).ColumnWidth
                        except:
                            pass
                    
                    temp_ws.Application.ActiveWindow.Zoom = 115
                    data_range = temp_ws.Range(temp_ws.Cells(1, 1), temp_ws.Cells(total_rows, capture_right))
                    data_range.CopyPicture(Appearance=1, Format=2)
                    time.sleep(0.8)
                    
                    safe_name = safe_filename(group_name)
                    img_path = os.path.join(output_folder, f"{safe_name}.png")
                    
                    img_saved = False
                    for attempt in range(3):
                        try:
                            img = ImageGrab.grabclipboard()
                            if img:
                                img.save(img_path, "PNG")
                                self.log(f"  ✅ {safe_name}.png")
                                img_saved = True
                                break
                            time.sleep(0.5)
                            data_range.CopyPicture(Appearance=1, Format=2)
                            time.sleep(0.5)
                        except:
                            time.sleep(0.5)
                    
                    if not img_saved:
                        img_path = ""
                    
                    excel.DisplayAlerts = False
                    try:
                        temp_ws.Delete()
                    except:
                        pass
                    excel.DisplayAlerts = True
                    
                    # Build DM list with dmNameRep from contacts
                    dms_for_tagging = []
                    for dm_name in all_dm_names:
                        dm_info = find_dm_info(dm_name, dm_map)
                        dms_for_tagging.append({
                            "name": dm_name,
                            "phone": dm_info['phone'],
                            "dmNameRep": dm_info['dmNameRep'],
                        })
                    
                    group_info_list.append({
                        "whatsapp_group": group_name,
                        "send_mode": send_mode,
                        "image": img_path,
                        "record_count": len(all_rows),
                        "markets": list(group_data['markets']),
                        "dms": dms_for_tagging,
                    })
                    
                else:
                    # ═══════════════════════════════════════════
                    #  ACCESSORIES MODE: One SS per DM
                    # ═══════════════════════════════════════════
                    
                    dms_for_tagging = []
                    dm_images = []
                    
                    for dm_name in all_dm_names:
                        ss_count += 1
                        rows = sorted(group_data['dm_data'][dm_name])
                        
                        # Get dmNameRep from contacts
                        dm_info = {
                            'phone': dm_map.get(dm_name.lower(), {}).get('phone') if dm_map else None,
                            'dmNameRep': get_dm_name_rep_fixed(dm_name)
                        }
                        rep_name = dm_info['dmNameRep']
                        
                        self.log(f"[{ss_count}] {dm_name} → @{rep_name} ({len(rows)} rows) → {group_name}")
                        self.progress_var.set(f"[{ss_count}] {dm_name}")
                        
                        # Create temp sheet
                        excel.DisplayAlerts = False
                        try:
                            wb.Sheets("_TEMP_SS").Delete()
                        except:
                            pass
                        excel.DisplayAlerts = True
                        
                        excel.DisplayAlerts = False
                        temp_ws = wb.Sheets.Add()
                        temp_ws.Name = "_TEMP_SS"
                        excel.DisplayAlerts = True
                        
                        # Copy header
                        ws.Range(ws.Cells(header_row, capture_left), ws.Cells(header_row, capture_right)).Copy(
                            temp_ws.Range(temp_ws.Cells(1, 1), temp_ws.Cells(1, capture_right)))
                        temp_ws.Rows(1).RowHeight = ws.Rows(header_row).RowHeight
                        
                        # Copy DM rows
                        for t_row, s_row in enumerate(rows, 2):
                            ws.Range(ws.Cells(s_row, capture_left), ws.Cells(s_row, capture_right)).Copy(
                                temp_ws.Range(temp_ws.Cells(t_row, 1), temp_ws.Cells(t_row, capture_right)))
                            try:
                                temp_ws.Rows(t_row).RowHeight = ws.Rows(s_row).RowHeight
                            except:
                                pass
                        
                        total_rows = len(rows) + 1
                        
                        for c in range(capture_left, capture_right + 1):
                            try:
                                temp_ws.Columns(c).ColumnWidth = ws.Columns(c).ColumnWidth
                            except:
                                pass
                        
                        temp_ws.Application.ActiveWindow.Zoom = 115
                        data_range = temp_ws.Range(temp_ws.Cells(1, 1), temp_ws.Cells(total_rows, capture_right))
                        data_range.CopyPicture(Appearance=1, Format=2)
                        time.sleep(0.8)
                        
                        safe_market = safe_filename(list(group_data['markets'])[0])
                        safe_dm = safe_filename(dm_name)
                        img_name = f"{safe_market}_{safe_dm}"
                        img_path = os.path.join(output_folder, f"{img_name}.png")
                        
                        img_saved = False
                        for attempt in range(3):
                            try:
                                img = ImageGrab.grabclipboard()
                                if img:
                                    img.save(img_path, "PNG")
                                    self.log(f"    ✅ {img_name}.png")
                                    img_saved = True
                                    break
                                time.sleep(0.5)
                                data_range.CopyPicture(Appearance=1, Format=2)
                                time.sleep(0.5)
                            except:
                                time.sleep(0.5)
                        
                        if not img_saved:
                            img_path = ""
                        
                        dm_images.append({"name": dm_name, "image": img_path})
                        dms_for_tagging.append({
                            "name": dm_name,
                            "phone": dm_info['phone'],
                            "dmNameRep": rep_name,
                        })
                        
                        excel.DisplayAlerts = False
                        try:
                            temp_ws.Delete()
                        except:
                            pass
                        excel.DisplayAlerts = True
                        time.sleep(0.2)
                    
                    group_info_list.append({
                        "whatsapp_group": group_name,
                        "send_mode": send_mode,
                        "markets": list(group_data['markets']),
                        "dms": dms_for_tagging,
                        "dm_images": dm_images,
                        "record_count": sum(len(group_data['dm_data'][d]) for d in all_dm_names),
                    })
                
                time.sleep(0.3)
            
            # Final cleanup
            excel.DisplayAlerts = False
            wb.Close(SaveChanges=False)
            excel.Quit()
            
            # Save metadata
            list_path = os.path.join(output_folder, "_group_list.json")
            with open(list_path, "w") as f:
                json.dump(group_info_list, f, indent=2, default=str)
            
            self.log(f"\n{'='*60}")
            self.log(f"✅ Done! {ss_count} screenshots for {len(group_map)} groups")
            self.log(f"📁 {output_folder}")
            self.log(f"📱 Ready for WhatsApp sender!")
            
        except Exception as e:
            self.log(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
        
        self.root.after(0, self.generation_done)
    
    def generation_done(self):
        self.generate_btn.configure(state=tk.NORMAL)
        self.progress_var.set("Done!")


def main():
    root = tk.Tk()
    app = ScreenshotGeneratorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()